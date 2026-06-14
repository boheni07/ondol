#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OnDol WBS 파생자산 생성기 (SSOT = docs/08-wbs-v1.1.md)
생성물:
  1. docs/wbs-github-issues.csv  — GitHub 이슈 임포트 CSV (작업패키지 1개 = 이슈 1개)
  2. docs/09-wbs-github.md        — GitHub 등록 가이드
  3. docs/08-wbs.xlsx             — WBS 시트 + Gantt_Chart 시트 (COUNTIF/SUMIF 연계)
  (옵션) scripts/import-wbs-github.sh 재생성은 09 가이드/CSV로 대체 가능하나 본 스크립트에서 함께 생성

설계: 본 md를 단일 원천으로 모든 작업 행을 파싱하고, 작업패키지(WBS 하위 섹션)별로
      그룹화하여 공수/작업수/우선순위를 집계한다. 이슈 단위 메타(일정 주차·라벨 type·담당 role)는
      PACKAGE_META 정적 표로 보강한다. 데이터 수치는 전적으로 md 파싱 결과를 따른다.
"""
import re, csv, sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
SSOT = ROOT / "docs" / "08-wbs-v1.1.md"
CSV_OUT = ROOT / "docs" / "wbs-github-issues.csv"
MD09_OUT = ROOT / "docs" / "09-wbs-github.md"
XLSX_OUT = ROOT / "docs" / "08-wbs.xlsx"
SH_OUT = ROOT / "scripts" / "import-wbs-github.sh"
SH_OUT = ROOT / "scripts" / "import-wbs-github.sh"

EFFORT_DAYS = {"XS": 0.5, "S": 1.0, "M": 2.5, "L": 5.0, "XL": 7.0}

# ── 작업패키지(이슈) 메타: (issue_key, 제목, Phase번호, Phase제목, 일정주차, 담당role, label type, milestone주차(due))
# issue_key 는 작업 ID 의 접두(파싱 시 매칭). 작업은 등장 순서대로 가장 긴 매칭 issue_key 에 귀속.
# Phase 11.5 는 별도 마일스톤. 5.G 신규 패키지, 17.4 신규 패키지.
PACKAGE_META = [
    # key,        title,                                            phase, phase_title,             weeks,    role,   ltype,     due_wk
    ("0.",        "[0] 프로젝트 셋업 + CI 기초 (Phase 0)",            "0",   "프로젝트 셋업 + CI 기초",  "W1–W2",  "PL",   "infra",   2),
    ("1.1.",      "[1.1] 마이그레이션 — 핵심 테이블",                  "1",   "DB 스키마 + RLS",         "W2–W4",  "PL",   "db",      4),
    ("1.2.",      "[1.2] RLS 정책",                                  "1",   "DB 스키마 + RLS",         "W2–W4",  "PL",   "db",      4),
    ("2.",        "[2] 인증 (Auth)",                                 "2",   "인증 (Auth)",             "W4–W6",  "Dev-A","backend", 6),
    ("3.1.",      "[3.1] 사용자 (users)",                            "3",   "사용자·당사자·매핑",       "W5–W7",  "Dev-A","backend", 7),
    ("3.2.",      "[3.2] 당사자 (persons)",                          "3",   "사용자·당사자·매핑",       "W5–W7",  "Dev-A","backend", 7),
    ("3.3.",      "[3.3] 보호자-당사자 매핑 (guardian_persons)",      "3",   "사용자·당사자·매핑",       "W5–W7",  "Dev-A","backend", 7),
    ("3.4.",      "[3.4] 당사자 계정 (person_accounts)",              "3",   "사용자·당사자·매핑",       "W5–W7",  "Dev-A","backend", 7),
    ("4.1.",      "[4.1] 권한 CRUD",                                 "4",   "권한 관리",               "W6–W8",  "Dev-A","backend", 8),
    ("4.2.",      "[4.2] 권한 위자드 플로우 (4 step)",                "4",   "권한 관리",               "W6–W8",  "Dev-A","backend", 8),
    ("4.3.",      "[4.3] 권한 변경 이력 (permission_logs)",           "4",   "권한 관리",               "W6–W8",  "Dev-A","backend", 8),
    ("4.4.",      "[4.4] RLS 권한 검증",                             "4",   "권한 관리",               "W6–W8",  "Dev-A","backend", 8),
    ("5.0.",      "[5.0] 기록 공통 CRUD",                            "5",   "기록 (Records)",          "W7–W12", "Dev-B","backend", 12),
    ("5.A.",      "[5.A] 의료 기록 (MED-001 ~ MED-010)",             "5",   "기록 (Records)",          "W7–W12", "Dev-B","backend", 12),
    ("5.B.",      "[5.B] 교육 기록 (EDU-001 ~ EDU-009)",             "5",   "기록 (Records)",          "W7–W12", "Dev-B","backend", 12),
    ("5.C.",      "[5.C] 복지 기록 (WEL-001 ~ WEL-007)",             "5",   "기록 (Records)",          "W7–W12", "Dev-B","backend", 12),
    ("5.D.",      "[5.D] 일상/돌봄 기록 (DAI-001 ~ DAI-005)",         "5",   "기록 (Records)",          "W7–W12", "Dev-B","backend", 12),
    ("5.E.",      "[5.E] 전환/자립 기록 (TRA-001 ~ TRA-007)",         "5",   "기록 (Records)",          "W7–W12", "Dev-B","backend", 12),
    ("5.F.",      "[5.F] 법적/행정 기록 (LEG-001 ~ LEG-005)",         "5",   "기록 (Records)",          "W7–W12", "Dev-B","backend", 12),
    ("5.G.",      "[5.G] 고유식별정보 암호화 서비스 (PIPA §24)",        "5",   "기록 (Records)",          "W7–W12", "Dev-B","security", 12),
    ("6.",        "[6] 자기표현 (Self-Expression)",                  "6",   "자기표현",                "W9–W10", "Dev-B","backend", 10),
    ("7.",        "[7] 파일 첨부 (Record Files)",                    "7",   "파일 첨부",               "W9–W10", "Dev-B","backend", 10),
    ("8.1.",      "[8.1] 이정표 CRUD",                               "8",   "이정표·타임라인",          "W10–W12","Dev-A","backend", 12),
    ("8.2.",      "[8.2] 타임라인 (records + milestones + self_expressions 통합)", "8", "이정표·타임라인", "W10–W12","Dev-A","backend", 12),
    ("9.1.",      "[9.1] 인수인계 CRUD",                             "9",   "인수인계",                "W11–W13","Dev-A","backend", 13),
    ("9.2.",      "[9.2] 인계 플로우",                               "9",   "인수인계",                "W11–W13","Dev-A","backend", 13),
    ("10.1.",     "[10.1] 알림 인프라",                              "10",  "알림",                    "W11–W13","Dev-A","backend", 13),
    ("10.2.",     "[10.2] 알림 채널",                                "10",  "알림",                    "W11–W13","Dev-A","backend", 13),
    ("11.5.",     "[11.5] 백엔드 통합 테스트 체크포인트",              "11.5","백엔드 통합 테스트 체크포인트","W13",  "PM",   "qa",      13),
    ("11.",       "[11] 접근 로그 (Audit)",                          "11",  "접근 로그",               "W12–W13","PL",   "security",13),
    ("12.1.",     "[12.1] 토큰 (07-design-system.md 기반)",          "12",  "디자인 시스템",            "W4–W6",  "Dev-C","design",  6),
    ("12.2.",     "[12.2] Primitives",                               "12",  "디자인 시스템",            "W4–W6",  "Dev-C","design",  6),
    ("12.3.",     "[12.3] Composite Components",                     "12",  "디자인 시스템",            "W4–W6",  "Dev-C","design",  6),
    ("12.4.",     "[12.4] 레이아웃 셀",                              "12",  "디자인 시스템",            "W4–W6",  "Dev-C","design",  6),
    ("13.1.",     "[13.1] 인증 화면",                                "13",  "웹 UI — 보호자",          "W7–W11", "Dev-C","frontend",11),
    ("13.2.",     "[13.2] 메인 화면",                                "13",  "웹 UI — 보호자",          "W7–W11", "Dev-C","frontend",11),
    ("13.3.",     "[13.3] 설정 화면",                                "13",  "웹 UI — 보호자",          "W7–W11", "Dev-C","frontend",11),
    ("14.",       "[14] 웹 UI — 당사자 (Person, 접근성 모드, 10 screens)","14","웹 UI — 당사자(접근성)","W10–W12","Dev-C","frontend",12),
    ("15.1.",     "[15.1] 활동지원사 (S)",                           "15",  "웹 UI — 전문가(4역할)",   "W9–W14", "Dev-D","frontend",14),
    ("15.2.",     "[15.2] 특수교사 (T)",                             "15",  "웹 UI — 전문가(4역할)",   "W9–W14", "Dev-D","frontend",14),
    ("15.3.",     "[15.3] 사회복지사 (W)",                           "15",  "웹 UI — 전문가(4역할)",   "W9–W14", "Dev-D","frontend",14),
    ("15.4.",     "[15.4] 치료사 (TH)",                              "15",  "웹 UI — 전문가(4역할)",   "W9–W14", "Dev-D","frontend",14),
    ("16.1.",     "[16.1] 인증",                                     "16",  "모바일 앱(RN)",           "W11–W15","Dev-E","mobile",  15),
    ("16.2.",     "[16.2] 보호자",                                   "16",  "모바일 앱(RN)",           "W11–W15","Dev-E","mobile",  15),
    ("16.3.",     "[16.3] 당사자 (접근성)",                          "16",  "모바일 앱(RN)",           "W11–W15","Dev-E","mobile",  15),
    ("16.4.",     "[16.4] 활동지원사",                               "16",  "모바일 앱(RN)",           "W11–W15","Dev-E","mobile",  15),
    ("16.5.",     "[16.5] 전문가 4역할 통합 패턴",                    "16",  "모바일 앱(RN)",           "W11–W15","Dev-E","mobile",  15),
    ("16.6.",     "[16.6] 모바일 전용 기능",                         "16",  "모바일 앱(RN)",           "W11–W15","Dev-E","mobile",  15),
    ("17.1.",     "[17.1] SEO (웹만)",                               "17",  "SEO·보안·접근성",         "W13–W15","PL",   "frontend",15),
    ("17.2.",     "[17.2] 보안",                                     "17",  "SEO·보안·접근성",         "W13–W15","PL",   "security",15),
    ("17.3.",     "[17.3] 접근성 (WCAG 2.1 AA)",                     "17",  "SEO·보안·접근성",         "W13–W15","PL",   "frontend",15),
    ("17.4.",     "[17.4] 법무·고지 (개인정보 거버넌스)",             "17",  "SEO·보안·접근성",         "W13–W15","PL",   "security",15),
    ("18.1.",     "[18.1] 자동화 테스트",                            "18",  "QA·테스트",               "W14–W16","PM",   "qa",      16),
    ("18.2.",     "[18.2] Zero Script QA (역할별 시나리오)",          "18",  "QA·테스트",               "W14–W16","PM",   "qa",      16),
    ("18.3.",     "[18.3] 성능",                                     "18",  "QA·테스트",               "W14–W16","PM",   "qa",      16),
    ("19.1.",     "[19.1] CI",                                       "19",  "CI/CD·배포",              "W15–W16","PL",   "infra",   16),
    ("19.2.",     "[19.2] CD",                                       "19",  "CI/CD·배포",              "W15–W16","PL",   "infra",   16),
    ("19.3.",     "[19.3] 운영",                                     "19",  "CI/CD·배포",              "W15–W16","PL",   "infra",   16),
]

ROLE_LABEL = {"PM": "PM", "PL": "PL", "Dev-A": "개발자 A", "Dev-B": "개발자 B",
              "Dev-C": "개발자 C", "Dev-D": "개발자 D", "Dev-E": "개발자 E"}
ROLE_FULL = {"PM": "PM (PM)", "PL": "PL (PL)", "Dev-A": "개발자 A (Dev-A)", "Dev-B": "개발자 B (Dev-B)",
             "Dev-C": "개발자 C (Dev-C)", "Dev-D": "개발자 D (Dev-D)", "Dev-E": "개발자 E (Dev-E)"}

# Phase 마일스톤 메타(20개; 11.5 포함 시 21개 마일스톤). due_wk·title 매핑.
MILESTONES = [
    ("Phase 0 · 프로젝트 셋업", "프로젝트 셋업 (W1-W2)", "W1–W2", 2),
    ("Phase 1 · DB 스키마 + RLS", "DB 스키마 + RLS (W2-W4)", "W2–W4", 4),
    ("Phase 2 · 인증 (Auth)", "인증 (Auth) (W4-W6)", "W4–W6", 6),
    ("Phase 3 · 사용자·당사자·매핑", "사용자·당사자·매핑 (W5-W7)", "W5–W7", 7),
    ("Phase 4 · 권한 관리", "권한 관리 (W6-W8)", "W6–W8", 8),
    ("Phase 5 · 기록 (Records)", "기록 (Records) (W7-W12)", "W7–W12", 12),
    ("Phase 6 · 자기표현", "자기표현 (W9-W10)", "W9–W10", 10),
    ("Phase 7 · 파일 첨부", "파일 첨부 (W9-W10)", "W9–W10", 10),
    ("Phase 8 · 이정표·타임라인", "이정표·타임라인 (W10-W12)", "W10–W12", 12),
    ("Phase 9 · 인수인계", "인수인계 (W11-W13)", "W11–W13", 13),
    ("Phase 10 · 알림", "알림 (W11-W13)", "W11–W13", 13),
    ("Phase 11 · 접근 로그", "접근 로그 (W12-W13)", "W12–W13", 13),
    ("Phase 11.5 · 백엔드 통합 테스트", "백엔드 통합 테스트 체크포인트 (W13)", "W13", 13),
    ("Phase 12 · 디자인 시스템", "디자인 시스템 (W4-W6)", "W4–W6", 6),
    ("Phase 13 · 웹 UI — 보호자", "웹 UI — 보호자 (W7-W11)", "W7–W11", 11),
    ("Phase 14 · 웹 UI — 당사자(접근성)", "웹 UI — 당사자(접근성) (W10-W12)", "W10–W12", 12),
    ("Phase 15 · 웹 UI — 전문가(4역할)", "웹 UI — 전문가(4역할) (W9-W14)", "W9–W14", 14),
    ("Phase 16 · 모바일 앱(RN)", "모바일 앱(RN) (W11-W15)", "W11–W15", 15),
    ("Phase 17 · SEO·보안·접근성", "SEO·보안·접근성 (W13-W15)", "W13–W15", 15),
    ("Phase 18 · QA·테스트", "QA·테스트 (W14-W16)", "W14–W16", 16),
    ("Phase 19 · CI/CD·배포", "CI/CD·배포 (W15-W16)", "W15–W16", 16),
]
# 마일스톤 제목 → 작업패키지 phase_title 매핑용
PHASE_MS_TITLE = {
    "0": "Phase 0 · 프로젝트 셋업", "1": "Phase 1 · DB 스키마 + RLS", "2": "Phase 2 · 인증 (Auth)",
    "3": "Phase 3 · 사용자·당사자·매핑", "4": "Phase 4 · 권한 관리", "5": "Phase 5 · 기록 (Records)",
    "6": "Phase 6 · 자기표현", "7": "Phase 7 · 파일 첨부", "8": "Phase 8 · 이정표·타임라인",
    "9": "Phase 9 · 인수인계", "10": "Phase 10 · 알림", "11": "Phase 11 · 접근 로그",
    "11.5": "Phase 11.5 · 백엔드 통합 테스트", "12": "Phase 12 · 디자인 시스템",
    "13": "Phase 13 · 웹 UI — 보호자", "14": "Phase 14 · 웹 UI — 당사자(접근성)",
    "15": "Phase 15 · 웹 UI — 전문가(4역할)", "16": "Phase 16 · 모바일 앱(RN)",
    "17": "Phase 17 · SEO·보안·접근성", "18": "Phase 18 · QA·테스트", "19": "Phase 19 · CI/CD·배포",
}

NEW_IDS = {"0.13", "0.14", "1.1.14", "1.1.15", "1.1.16", "1.2.6", "1.2.7", "2.12", "2.13",
           "3.2.10", "5.G.1", "11.5.1", "11.5.2", "11.5.3", "11.9", "11.10",
           "13.1.7", "13.1.8", "13.3.6", "17.2.9", "17.4.1"}

ID_RE = re.compile(r"^\|\s*([0-9]+(?:\.[0-9A-G]+)*)\s*(?:✨\s*NEW)?\s*\|")
EFFORT_RE = re.compile(r"\b(XS|S|M|L|XL)\b")
PRIO_RE = re.compile(r"\b(P[012])\b")


def parse_tasks():
    """SSOT md 의 모든 작업 테이블 행을 파싱해 (id, title, effort, prio, assignee_role, detail) 리스트 반환."""
    lines = SSOT.read_text(encoding="utf-8").splitlines()
    tasks = []
    for ln in lines:
        m = ID_RE.match(ln)
        if not m:
            continue
        tid = m.group(1)
        # 셀 분해
        cells = [c.strip() for c in ln.strip().strip("|").split("|")]
        # 첫 셀은 ID(✨ NEW 포함). 마지막 셀은 담당자.
        first = cells[0]
        # ID 정규화 (✨ NEW 제거)
        idtxt = first.replace("✨", "").replace("NEW", "").strip()
        if idtxt != tid:
            tid = re.split(r"\s", idtxt)[0]
        title = cells[1] if len(cells) > 1 else ""
        assignee_full = cells[-1]
        # 공수/우선순위: 셀 중 정확히 EFFORT/PRIO 인 셀을 탐색
        effort, prio = None, None
        for c in cells:
            if c in EFFORT_DAYS and effort is None:
                effort = c
            if c in ("P0", "P1", "P2") and prio is None:
                prio = c
        if effort is None or prio is None:
            continue  # 작업 행이 아님(요약표 등)
        # 세부사항: 가장 긴 한글 설명 셀(참고 포함). 보통 prio 다음 셀.
        detail = ""
        try:
            pidx = cells.index(prio)
            # detail 은 prio 다음, AI 프롬프트(`...`) 이전 셀
            cand = cells[pidx + 1] if pidx + 1 < len(cells) else ""
            detail = cand
        except ValueError:
            pass
        role = full_to_role(assignee_full)
        tasks.append(dict(id=tid, title=title, effort=effort, prio=prio,
                          role=role, detail=detail, new=(tid in NEW_IDS)))
    return tasks


def full_to_role(full):
    full = full.strip()
    mapping = {"PM": "PM", "PL": "PL", "개발자 A": "Dev-A", "개발자 B": "Dev-B",
               "개발자 C": "Dev-C", "개발자 D": "Dev-D", "개발자 E": "Dev-E"}
    return mapping.get(full, full)


def assign_package(tid):
    """가장 긴 prefix 매칭으로 작업패키지 결정."""
    best = None
    for meta in PACKAGE_META:
        key = meta[0]
        if tid.startswith(key):
            if best is None or len(key) > len(best[0]):
                best = meta
    return best


def build_packages(tasks):
    pkgs = {}  # title -> dict
    order = []
    for t in tasks:
        meta = assign_package(t["id"])
        if meta is None:
            print(f"WARN: 작업패키지 미매칭 {t['id']}", file=sys.stderr)
            continue
        key, title, phase, ptitle, weeks, role, ltype, due = meta
        if title not in pkgs:
            pkgs[title] = dict(title=title, phase=phase, ptitle=ptitle, weeks=weeks,
                               role=role, ltype=ltype, due=due, tasks=[])
            order.append(title)
        pkgs[title]["tasks"].append(t)
    for title in order:
        p = pkgs[title]
        p["count"] = len(p["tasks"])
        p["days"] = round(sum(EFFORT_DAYS[t["effort"]] for t in p["tasks"]), 1)
        # 패키지 우선순위 = 최고 우선(P0>P1>P2)
        prios = [t["prio"] for t in p["tasks"]]
        p["prio"] = "P0" if "P0" in prios else ("P1" if "P1" in prios else "P2")
    return [pkgs[t] for t in order]


def fmt_days(d):
    return (f"{d:.1f}".rstrip("0").rstrip(".")) + "d"


def body_for(pkg):
    role = pkg["role"]
    lines = [
        f"**Phase:** {pkg['phase']} · {pkg['ptitle']}  ",
        f"**담당자:** {ROLE_FULL[role]}  ",
        f"**우선순위:** {pkg['prio']}  ·  **예상 공수:** {fmt_days(pkg['days'])}  ·  "
        f"**작업 수:** {pkg['count']}  ·  **일정:** {pkg['weeks']}",
        "",
        "### 작업 목록",
    ]
    for t in pkg["tasks"]:
        nm = " ✨ NEW" if t["new"] else ""
        det = (" — " + t["detail"]) if t["detail"] else ""
        lines.append(f"- [ ] **{t['id']}**{nm} {t['title']} `{t['effort']}` `{t['prio']}`{det}")
    lines.append("")
    lines.append(f"> 원본: docs/08-wbs-v1.1.md · {pkg['title'][pkg['title'].find(']')+2:]}")
    return "\n".join(lines)


def write_csv(pkgs):
    with CSV_OUT.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        w.writerow(["Title", "Body", "Milestone", "Labels", "Assignee(role)",
                    "Priority", "Estimate(days)", "Phase", "WorkPackage"])
        for p in pkgs:
            ms = PHASE_MS_TITLE[p["phase"]]
            labels = f"phase:{p['phase']},priority:{p['prio']},type:{p['ltype']},role:{p['role']}"
            wp = p["title"][p["title"].find("]") + 2:]
            est = (f"{p['days']:.1f}".rstrip("0").rstrip("."))
            w.writerow([p["title"], body_for(p), ms, labels, ROLE_LABEL[p["role"]],
                        p["prio"], est, p["phase"], wp])


def write_md09(pkgs):
    # 담당자별 집계
    roles = ["PM", "PL", "Dev-A", "Dev-B", "Dev-C", "Dev-D", "Dev-E"]
    agg = {r: dict(phases=set(), pkgs=0, tasks=0, days=0.0) for r in roles}
    for p in pkgs:
        a = agg[p["role"]]
        a["phases"].add(p["phase"])
        a["pkgs"] += 1
        a["tasks"] += p["count"]
        a["days"] += p["days"]

    role_area = {
        "PM": "QA·테스트(품질/인수 시나리오) + 백엔드 통합 테스트 체크포인트",
        "PL": "셋업·CI·DB·감사·보안접근성·법무고지·배포",
        "Dev-A": "백엔드(인증·사용자·권한·타임라인·인계·알림·동의)",
        "Dev-B": "백엔드(기록·자기표현·파일·고유식별정보 암호화)",
        "Dev-C": "프론트 웹(디자인시스템·보호자·당사자·동의/권리)",
        "Dev-D": "프론트 웹(전문가 4역할)",
        "Dev-E": "모바일(RN)",
    }

    def phs(s):
        # 정렬: 숫자(11.5 포함)
        return ",".join(sorted(s, key=lambda x: float(x)))

    out = []
    out.append("# WBS → GitHub 등록 가이드 (Issues · Milestones · Projects)")
    out.append("")
    out.append("> 생성: project-planning 스킬 · 원본 docs/08-wbs-v1.1.md (정본 SSOT, v1.1)")
    out.append("")
    out.append(f"작업패키지(WBS 하위 섹션) 1개 = Issue 1개, Phase 1개 = Milestone 1개, "
               f"전체를 Projects 보드 1개로 관리한다. v1.1 기준 **이슈 {len(pkgs)}개 · "
               f"마일스톤 {len(MILESTONES)}개 · 작업 {sum(p['count'] for p in pkgs)}개 · "
               f"공수 {fmt_days(round(sum(p['days'] for p in pkgs),1))}**.")
    out.append("")
    out.append("---")
    out.append("")
    out.append("## 1. 담당자 배정 정책")
    out.append("")
    out.append("| 담당자 | 담당 영역 | Phase | 작업패키지 | 작업 | 예상 공수 |")
    out.append("|--------|-----------|-------|:----:|:---:|:------:|")
    for r in roles:
        a = agg[r]
        out.append(f"| {ROLE_FULL[r]} | {role_area[r]} | {phs(a['phases'])} | "
                   f"{a['pkgs']} | {a['tasks']} | {fmt_days(round(a['days'],1))} |")
    out.append("")
    out.append("> 배정 근거: docs/08-wbs-v1.1.md 본문 담당자 열. 백엔드 부하 완화를 위해 "
               "타임라인·인계·알림(8~10)은 Dev-A, 감사·보안(11·17)은 PL, 백엔드 통합 테스트(11.5)는 PM이 담당. "
               "GitHub 실제 할당은 스크립트 상단 role→username 매핑 사용.")
    out.append("")
    out.append("---")
    out.append("")
    out.append(f"## 2. Milestones (Phase {len(MILESTONES)}개)")
    out.append("")
    out.append("| Milestone | 주차 | 작업패키지 | 작업 | 공수 | 담당(주) |")
    out.append("|-----------|:----:|:----:|:---:|:---:|:------:|")
    # phase별 집계
    for ms_title, _desc, weeks, _due in MILESTONES:
        phase = [k for k, v in PHASE_MS_TITLE.items() if v == ms_title][0]
        ph_pkgs = [p for p in pkgs if p["phase"] == phase]
        npk = len(ph_pkgs)
        ntask = sum(p["count"] for p in ph_pkgs)
        ndays = round(sum(p["days"] for p in ph_pkgs), 1)
        # 대표 담당 = 최다 패키지 role
        from collections import Counter
        rc = Counter(p["role"] for p in ph_pkgs)
        lead = ROLE_LABEL[rc.most_common(1)[0][0]] if rc else ""
        out.append(f"| {ms_title} | {weeks} | {npk} | {ntask} | {fmt_days(ndays)} | {lead} |")
    out.append("")
    out.append("---")
    out.append("")
    out.append("## 3. Labels")
    out.append("")
    out.append("- `phase:0` ~ `phase:19` (`phase:11.5` 포함)")
    out.append("- `priority:P0` · `priority:P1` · `priority:P2`")
    out.append("- `type:infra|db|backend|frontend|mobile|design|security|qa`")
    out.append("- `role:PM|PL|Dev-A|Dev-B|Dev-C|Dev-D|Dev-E`")
    out.append("")
    out.append("---")
    out.append("")
    out.append(f"## 4. Issue 목록 (작업패키지 {len(pkgs)}개)")
    out.append("")
    out.append("| # | Issue 제목 | Milestone | 담당자 | 우선 | 공수 | 작업수 |")
    out.append("|:--:|-----------|-----------|:------:|:---:|:---:|:---:|")
    for i, p in enumerate(pkgs, 1):
        ms = PHASE_MS_TITLE[p["phase"]]
        out.append(f"| {i} | {p['title'].replace('(Phase 0)','(Phase 0)')} | {ms} | "
                   f"{ROLE_LABEL[p['role']]} | {p['prio']} | {fmt_days(p['days'])} | {p['count']} |")
    out.append("")
    out.append("---")
    out.append("")
    out.append("## 5. 등록 방법")
    out.append("")
    out.append("### 방법 A — 자동 스크립트 (권장)")
    out.append("")
    out.append("```bash")
    out.append("gh auth login")
    out.append("gh auth refresh -s project,repo")
    out.append("# scripts/import-wbs-github.sh 상단 role→username 매핑 편집 후")
    out.append("bash scripts/import-wbs-github.sh")
    out.append("```")
    out.append("")
    out.append(f"스크립트는 라벨→마일스톤→이슈 {len(pkgs)}개→(옵션)Projects 순으로 동작, 멱등 재실행 가능.")
    out.append("")
    out.append("### 방법 B — CSV 임포트")
    out.append("")
    out.append("docs/wbs-github-issues.csv 를 GitHub CSV 임포트 도구로 업로드.")
    out.append("")
    out.append("### 방법 C — Projects 커스텀 필드")
    out.append("")
    out.append("Status·담당자·우선순위·공수(일)·Phase 필드 권장. 라벨로도 필터 가능.")
    out.append("")
    out.append("---")
    out.append("")
    out.append("> 재생성: `python scripts/gen-wbs-assets.py` (SSOT=docs/08-wbs-v1.1.md). "
               "csv·09·xlsx 세 자산을 일괄 갱신.")
    out.append("")
    MD09_OUT.write_text("\n".join(out), encoding="utf-8")


def write_xlsx(pkgs, tasks):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "WBS"

    hdr_fill = PatternFill("solid", fgColor="305496")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center")
    new_fill = PatternFill("solid", fgColor="FFF2CC")

    headers = ["Phase", "작업패키지", "ID", "작업", "공수(코드)", "공수(일)",
               "우선순위", "담당자", "NEW", "세부사항"]
    ws.append(headers)
    for c, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = border
    ws.freeze_panes = "A2"

    # 작업 순서대로 출력(패키지 메타로 phase/wp 부여)
    for t in tasks:
        meta = assign_package(t["id"])
        if meta is None:
            continue
        _, title, phase, ptitle, *_ = meta
        wp = title[title.find("]") + 2:]
        days = EFFORT_DAYS[t["effort"]]
        row = [phase, wp, t["id"], t["title"], t["effort"], days,
               t["prio"], ROLE_LABEL.get(t["role"], t["role"]),
               "NEW" if t["new"] else "", t["detail"]]
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).alignment = wrap if c == 10 else (center if c in (1,3,5,6,7,8,9) else wrap)
        if t["new"]:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = new_fill

    widths = [8, 34, 10, 40, 10, 9, 10, 12, 7, 70]
    for i, wdt in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt

    last = ws.max_row
    # 합계 행
    ws.append(["", "합계", "", "", "", f"=SUM(F2:F{last})", "", "", "", ""])
    tot_r = ws.max_row
    ws.cell(row=tot_r, column=2).font = Font(bold=True)
    ws.cell(row=tot_r, column=6).font = Font(bold=True)

    # ── Gantt_Chart 시트 (Phase별 COUNTIF/SUMIF 집계) ──
    gs = wb.create_sheet("Gantt_Chart")
    g_headers = ["Phase", "Phase명", "작업 수", "공수(일)", "MVP(P0) 작업 수",
                 "시작주", "종료주", "타임라인 (W1 ~ W16)"]
    gs.append(g_headers)
    for c, _ in enumerate(g_headers, 1):
        cell = gs.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = border

    # Phase 행 정의 (11.5 포함 21행)
    phase_rows = [
        ("0", "프로젝트 셋업 + CI 기초", 1, 2),
        ("1", "DB 스키마 + RLS", 2, 4),
        ("2", "인증 (Auth)", 4, 6),
        ("3", "사용자·당사자·매핑", 5, 7),
        ("4", "권한 관리", 6, 8),
        ("5", "기록 (Records)", 7, 12),
        ("6", "자기표현", 9, 10),
        ("7", "파일 첨부", 9, 10),
        ("8", "이정표·타임라인", 10, 12),
        ("9", "인수인계", 11, 13),
        ("10", "알림", 11, 13),
        ("11", "접근 로그", 12, 13),
        ("11.5", "백엔드 통합 테스트", 13, 13),
        ("12", "디자인 시스템", 4, 6),
        ("13", "웹 UI — 보호자", 7, 11),
        ("14", "웹 UI — 당사자(접근성)", 10, 12),
        ("15", "웹 UI — 전문가(4역할)", 9, 14),
        ("16", "모바일 앱(RN)", 11, 15),
        ("17", "SEO·보안·접근성", 13, 15),
        ("18", "QA·테스트", 14, 16),
        ("19", "CI/CD·배포", 15, 16),
    ]
    bar_fill = PatternFill("solid", fgColor="4472C4")
    g_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    n_weeks = 16
    timeline_start_col = 9  # I열부터 W1
    # 주차 헤더(2행)
    gs.cell(row=1, column=8, value="타임라인 (W1 ~ W16) →")
    for w in range(1, n_weeks + 1):
        cell = gs.cell(row=2, column=timeline_start_col - 1 + w, value=f"W{w}")
        cell.alignment = center
        cell.font = Font(size=8, bold=True)
        cell.border = g_thin
    gs.cell(row=2, column=1, value="").border = g_thin

    rownum = 3
    for ph, name, sw, ew in phase_rows:
        # WBS 시트의 Phase 열(A)이 텍스트이므로 COUNTIF/SUMIF 로 집계
        cnt = f'=COUNTIF(WBS!$A$2:$A${last},$A{rownum})'
        days = f'=SUMIF(WBS!$A$2:$A${last},$A{rownum},WBS!$F$2:$F${last})'
        # MVP(P0): Phase 일치 & 우선순위 P0 → COUNTIFS
        mvp = f'=COUNTIFS(WBS!$A$2:$A${last},$A{rownum},WBS!$G$2:$G${last},"P0")'
        gs.append([ph, name, cnt, days, mvp, sw, ew] + [""] * n_weeks)
        for c in range(1, 8):
            gs.cell(row=rownum, column=c).border = g_thin
            if c in (1, 3, 4, 5, 6, 7):
                gs.cell(row=rownum, column=c).alignment = center
        # 간트 막대
        for w in range(1, n_weeks + 1):
            cell = gs.cell(row=rownum, column=timeline_start_col - 1 + w)
            cell.border = g_thin
            if sw <= w <= ew:
                cell.fill = bar_fill
        rownum += 1

    # 합계 행
    tot = rownum
    gs.cell(row=tot, column=2, value="합계").font = Font(bold=True)
    gs.cell(row=tot, column=3, value=f"=SUM(C3:C{rownum-1})").font = Font(bold=True)
    gs.cell(row=tot, column=4, value=f"=SUM(D3:D{rownum-1})").font = Font(bold=True)
    gs.cell(row=tot, column=5, value=f"=SUM(E3:E{rownum-1})").font = Font(bold=True)
    for c in range(1, 8):
        gs.cell(row=tot, column=c).border = g_thin

    gwidths = [8, 26, 9, 10, 16, 8, 8]
    for i, wdt in enumerate(gwidths, 1):
        gs.column_dimensions[get_column_letter(i)].width = wdt
    for w in range(1, n_weeks + 1):
        gs.column_dimensions[get_column_letter(timeline_start_col - 1 + w)].width = 4
    gs.freeze_panes = "B3"

    wb.save(XLSX_OUT)


def write_sh(pkgs):
    """gh CLI 자동 등록 스크립트(라벨·마일스톤·이슈 58개) 재생성. csv/09 와 동일 데이터원."""
    L = []
    A = L.append
    A("#!/usr/bin/env bash")
    A("# OnDol WBS -> GitHub 자동 등록 (라벨/마일스톤/이슈)")
    A("# 생성: scripts/gen-wbs-assets.py (SSOT=docs/08-wbs-v1.1.md) · 멱등 재실행 가능")
    A("# 사용: 아래 GH_* 에 GitHub username 매핑 후 `bash scripts/import-wbs-github.sh`")
    A("set -uo pipefail")
    A('REPO="${REPO:-boheni07/ondol}"')
    A('START_DATE="${START_DATE:-2026-06-16}"')
    A("")
    A('GH_PM=""')
    A('GH_PL=""')
    for d in "ABCDE":
        A(f'GH_DEV_{d}=""')
    A("")
    A('role_user() { case "$1" in')
    A('  PM) echo "$GH_PM";; PL) echo "$GH_PL";; Dev-A) echo "$GH_DEV_A";; Dev-B) echo "$GH_DEV_B";;')
    A('  Dev-C) echo "$GH_DEV_C";; Dev-D) echo "$GH_DEV_D";; Dev-E) echo "$GH_DEV_E";; *) echo "";; esac; }')
    A('due_for() { date -u -d "$START_DATE +$(( $1 * 7 )) days" +%Y-%m-%dT%H:%M:%SZ 2>/dev/null || echo ""; }')
    A("")
    # 1) 라벨
    A('echo "== 1) 라벨 =="')
    phases = []
    types = []
    for m in PACKAGE_META:
        if m[2] not in phases:
            phases.append(m[2])
        if m[6] not in types:
            types.append(m[6])
    for ph in phases:
        A(f'gh label create "phase:{ph}" --color C5DEF5 --description "Phase {ph}" -R "$REPO" --force >/dev/null 2>&1 || true')
    for pr, col in (("P0", "B60205"), ("P1", "FBCA04"), ("P2", "0E8A16")):
        A(f'gh label create "priority:{pr}" --color {col} --description "{pr}" -R "$REPO" --force >/dev/null 2>&1 || true')
    for t in types:
        A(f'gh label create "type:{t}" --color D4C5F9 --description "{t}" -R "$REPO" --force >/dev/null 2>&1 || true')
    for r in ("PM", "PL", "Dev-A", "Dev-B", "Dev-C", "Dev-D", "Dev-E"):
        A(f'gh label create "role:{r}" --color BFD4F2 --description "{r}" -R "$REPO" --force >/dev/null 2>&1 || true')
    A("")
    # 2) 마일스톤
    A('echo "== 2) 마일스톤 =="')
    for title, desc, weeks, due_wk in MILESTONES:
        A(f"DUE=$(due_for {due_wk})")
        A(f'gh api -X POST "repos/$REPO/milestones" -f title="{title}" -f state="open" '
          f'-f description="{desc}" ${{DUE:+-f due_on="$DUE"}} >/dev/null 2>&1 || true')
    A("")
    # 3) 이슈
    A(f'echo "== 3) 이슈 {len(pkgs)}개 =="')
    A("CREATED_URLS=()")
    for p in pkgs:
        ms = PHASE_MS_TITLE[p["phase"]]
        labels = f"phase:{p['phase']},priority:{p['prio']},type:{p['ltype']},role:{p['role']}"
        A(f"ASSIGNEE=$(role_user {p['role']})")
        A("read -r -d '' BODY <<'ONDOL_BODY_EOF' || true")
        for bl in body_for(p).split("\n"):
            A(bl)
        A("ONDOL_BODY_EOF")
        A(f'URL=$(gh issue create -R "$REPO" --title "{p["title"]}" --body "$BODY" '
          f'--milestone "{ms}" --label "{labels}" ${{ASSIGNEE:+--assignee "$ASSIGNEE"}} 2>/dev/null)')
        A(f'echo "  created: ${{URL:-(실패: {p["title"]})}}"')
        A('[ -n "${URL:-}" ] && CREATED_URLS+=("$URL")')
        A("")
    A(f'echo "== 완료: 이슈 {len(pkgs)}개 시도 =="')
    SH_OUT.write_text("\n".join(L) + "\n", encoding="utf-8")


def main():
    tasks = parse_tasks()
    pkgs = build_packages(tasks)
    total_tasks = sum(p["count"] for p in pkgs)
    total_days = round(sum(p["days"] for p in pkgs), 1)
    print(f"파싱: 작업 {len(tasks)}개, 패키지(이슈) {len(pkgs)}개, "
          f"집계 작업 {total_tasks}개, 공수 {total_days}d", file=sys.stderr)
    new_cnt = sum(1 for t in tasks if t["new"])
    print(f"NEW 작업: {new_cnt}개", file=sys.stderr)
    write_csv(pkgs)
    write_md09(pkgs)
    write_xlsx(pkgs, tasks)
    write_sh(pkgs)
    print("생성 완료: docs/wbs-github-issues.csv, docs/09-wbs-github.md, docs/08-wbs.xlsx, scripts/import-wbs-github.sh", file=sys.stderr)


if __name__ == "__main__":
    main()
